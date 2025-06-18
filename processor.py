import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl import Workbook
import win32com.client as win32
from openpyxl.styles import Font, Alignment, Border, Side
import xlsxwriter
import BCS_connector
from datetime import datetime
import matplotlib.pyplot as plt
import mailer



def read_data():

    ids = []

    transfer_df = BCS_connector.reader_df(BCS_connector.transfers)
    transfer_df['transfer_date'] = pd.to_datetime(transfer_df['transfer_date']).dt.date

    # Convert today's date to datetime.date
    today = pd.to_datetime('today').date()
    transfer_df['Days Since Created'] = (today - transfer_df['transfer_date']).apply(lambda x: x.days)
    transfer_df["transfer_tracking_no"] = transfer_df["transfer_tracking_no"].fillna("Not available (blank)")

    
    rma_df = BCS_connector.reader_df(BCS_connector.rma_created_not_recorded)
    ir_df = BCS_connector.reader_df(BCS_connector.ir_created_not_shipped)

 
    rma_df["order_date"] = pd.to_datetime(rma_df["order_date"])
    rma_df["order_date (year)"] = rma_df["order_date"].dt.year

    ir_df['date_created'] = pd.to_datetime(ir_df['date_created'])
    ir_df["date_created (year)"] = ir_df["date_created"].dt.year

    return transfer_df, rma_df, ir_df
    

from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl import load_workbook
def pivot_table(transfer_df, rma_df, ir_df):
    output_file = "D:\\Brians_report_automation\\Transfers and RMA\\wednesday_RMA_reports.xlsx"
    
    # Writing to Excel with different sheet names
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write dataframes to their respective sheets
        transfer_df.to_excel(writer, sheet_name='Transfers', index=False)
        rma_df.to_excel(writer, sheet_name='RMA created not Rec', index=False)
        ir_df.to_excel(writer, sheet_name = 'IR created but not shipped', index = False)

        transfer_df = transfer_df[transfer_df["shipped_flag"] == "Y"]
        transfer_df = transfer_df[transfer_df["Days Since Created"] >= 7]

        partial_transfer_df = transfer_df[transfer_df['received_date'].notnull()]

        # Create a pivot table from the rma_df
        transfers_pivot_df = pd.pivot_table(
            transfer_df, 
            values=['Days Since Created'], 
            index=['transfer_no', "from_name", "to_name", "shipped_flag", "transfer_tracking_no"], 
            aggfunc={'Days Since Created': 'sum'},
            fill_value=0
        )

        # Create a pivot table from the rma_df
        partial_pivot_df = pd.pivot_table(
            partial_transfer_df, 
            values=['received_date'], 
            index=['transfer_no', "from_name", "to_name"],
            aggfunc='count', 
            fill_value=0
        )

        transfers_pivot_df_copied = transfers_pivot_df.copy()

        transfers_table = transfers_pivot_df_copied.to_html(classes='table table-striped', border=0, index=False)
        partial_transfer_table = partial_transfer_df.to_html(classes = 'table table-striped', border=0, index=False)

        # Add custom styling (bold column headers and outer borders)
        transfers_table = transfers_table.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
        transfers_table = transfers_table.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
        transfers_table = transfers_table.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')

        
        # Write the pivot table to a new sheet
        transfers_pivot_df.to_excel(writer, sheet_name='Pivot Summary - transfers')
        partial_pivot_df.to_excel(writer, sheet_name="Pivot Summary - Partial")

        # Create a pivot table from the rma_df
        rma_pivot_df = pd.pivot_table(
            rma_df, 
            values=['RMA_Number', 'open_line_value'], 
            index=['sales_location_name'], 
            columns=['order_date (year)'], 
            aggfunc={
                'RMA_Number': pd.Series.nunique,  # Distinct count for RMA_Number
                'open_line_value': 'sum'  # Sum for open_line_value
            },
            fill_value=0
        )

        rma_pivot_df['open_line_value'] = rma_pivot_df['open_line_value'].round(0).astype(int)

        # Write the pivot table to a new sheet
        rma_pivot_df.to_excel(writer, sheet_name='Pivot Summary - RMA')
        rma_table = rma_pivot_df.to_html(classes='table table-striped', border=1, index=False)

        # Add custom styling (bold column headers and outer borders)
        rma_table = rma_table.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
        rma_table = rma_table.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
        rma_table = rma_table.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')


        
        # create pivot for IR table
        ir_pivot_df = pd.pivot_table(
            ir_df, 
            values=['inventory_return_hdr_uid', 'l_unit_price'], 
            index=['location_name'], 
            columns=['date_created (year)'], 
            aggfunc={
                'inventory_return_hdr_uid': pd.Series.nunique,  # Distinct count of 'name'
                'l_unit_price': 'sum'  # Sum of 'amt'
            },
            fill_value=0
        )

        ir_pivot_df['l_unit_price'] = ir_pivot_df['l_unit_price'].round(0).astype(int)
        ir_pivot_df.to_excel(writer, sheet_name='Pivot Summary - IR')


    # Load the workbook to apply formatting
    wb = load_workbook(output_file)
    ws = wb['Pivot Summary - RMA']


    # Swap the first and second rows, avoiding merged cells
    for col_idx in range(1, len(ws[1]) + 1):  # Loop over columns in the first row
        cell1 = ws.cell(row=1, column=col_idx)
        cell2 = ws.cell(row=2, column=col_idx)

        # Check if the cell is part of a merged range
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            if min_row <= 1 <= max_row and min_col <= col_idx <= max_col:
                is_merged = True
                break

        # If not merged, swap the values
        if not is_merged:
            cell1.value, cell2.value = cell2.value, cell1.value

    # Create a border style for the header and title
    border = Border(
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Apply border and bold text to the new title row (first row, originally second)
    for cell in ws[1]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # Apply border and bold text to the new second row (originally first)
    for cell in ws[2]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # Apply border and bold text to the third row
    for cell in ws[3]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # Ensure no borders or bold font for the data rows
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = None  # Remove borders
            cell.font = Font(bold=False)  # Ensure the font is not bold

    amt_column_index = 3 

    # Apply number formatting with commas for the 'amt' column
    for row in ws.iter_rows(min_row=4, min_col=amt_column_index, max_col=amt_column_index):
        for cell in row:
            cell.number_format = '#,##0'  # Comma separator for thousands

    # Set column widths based on the content
    for col in ws.columns:
        max_length = 0
        first_cell = col[0]  # Get the first cell in the column

        # Check if the first cell in the column is part of a merged range
        if any(first_cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges):
            # If the first cell is merged, skip it and find the first non-merged cell
            first_cell = next((cell for cell in col if not any(cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges)), first_cell)

        # Now, we have a valid first cell (not merged), so we can safely access column_letter
        column = first_cell.column_letter  # Get the column name (column letter)

        # Iterate over all cells in the column to find the maximum length of values
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Set the adjusted width (with a small buffer of +2)
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook with applied formatting
    wb.save(output_file)
    
    # Load the workbook to apply formatting
    wb = load_workbook(output_file)
    ws = wb['Pivot Summary - transfers']

    # Create a border style for the header and title
    border = Border(
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Apply border and bold text to the title row (first row)
    for cell in ws[1]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)


    # Ensure no borders or bold font for the data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = None  # Remove borders
            cell.font = Font(bold=False)  # Ensure the font is not bold

    # Set column widths based on the content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


    # Save the workbook with applied formatting
    wb.save(output_file)

    wb = load_workbook(output_file)
    ws = wb['Pivot Summary - Partial']

    # Create a border style for the header and title
    border = Border(
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Apply border and bold text to the title row (first row)
    for cell in ws[1]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)


    # Ensure no borders or bold font for the data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = None  # Remove borders
            cell.font = Font(bold=False)  # Ensure the font is not bold

    # Set column widths based on the content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


    # Save the workbook with applied formatting
    wb.save(output_file)




    # Load the workbook to apply formatting
    wb = load_workbook(output_file)
    ws = wb['Pivot Summary - IR']

    # Create a border style for the header and title
    border = Border(
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Apply border and bold text to the title row (first row)
    for cell in ws[1]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    for cell in ws[2]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    for cell in ws[3]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)



    # Ensure no borders or bold font for the data rows
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = None  # Remove borders
            cell.font = Font(bold=False)  # Ensure the font is not bold


    ir_amt_column_index_1 = 5 
    ir_amt_column_index_2 = 7

    # Apply number formatting with commas for the 'amt' column
    for row in ws.iter_rows(min_row=4, min_col=ir_amt_column_index_1, max_col=ir_amt_column_index_1):
        for cell in row:
            cell.number_format = '#,##0'  # Comma separator for thousands

    # Apply number formatting with commas for the 'amt' column
    for row in ws.iter_rows(min_row=4, min_col=ir_amt_column_index_2, max_col=ir_amt_column_index_2):
        for cell in row:
            cell.number_format = '#,##0'  # Comma separator for thousands

    # Set column widths based on the content
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name using get_column_letter
        for cell in col:
            # Ensure we're working with a non-merged cell
            if cell.coordinate == cell.parent.merged_cells:  # Skip merged cells
                continue
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass


    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

    # Save the workbook with applied formatting
    wb.save(output_file)

    print(f'Pivot table created and saved to {output_file}')

    df = pd.read_excel(output_file, sheet_name="Pivot Summary - transfers", index_col=None)
    df_pt = pd.read_excel(output_file, sheet_name="Pivot Summary - Partial", index_col=None)
    transfers_table = df.to_html(classes='table table-striped', border=1, index=False)
    partial_transfer_table = df_pt.to_html(classes='table table-striped', border=1, index=False)


    # Add custom styling (bold column headers and outer borders)
    transfers_table = transfers_table.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
    transfers_table = transfers_table.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
    transfers_table = transfers_table.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')


    # Add custom styling (bold column headers and outer borders)
    partial_transfer_table = partial_transfer_table.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
    partial_transfer_table = partial_transfer_table.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
    partial_transfer_table = partial_transfer_table.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')

    df = pd.read_excel(output_file, sheet_name="Pivot Summary - IR")

    ir_table = ir_pivot_df.to_html(classes='table table-striped', border=0, index=False)
    # Add custom styling (bold column headers and outer borders)
    # Add custom styling (bold column headers and outer borders)
    ir_table = ir_table.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
    ir_table = ir_table.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
    ir_table = ir_table.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')


    return transfers_table, rma_table, ir_table, partial_transfer_table



def main():
    tdf, rdf, irdf = read_data()
    transfers_table, rma_table, ir_table, partial = pivot_table(tdf, rdf, irdf)
    mailer.sender("D:\\Brians_report_automation\\Transfers and RMA\\wednesday_RMA_reports.xlsx", transfers_table, rma_table, ir_table, partial)


if __name__ == "__main__":
    main()


"""
"""