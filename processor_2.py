import pandas as pd
import openpyxl
import numpy as np
from openpyxl import Workbook
import win32com.client as win32
from openpyxl.styles import Font, Alignment, Border, Side
import xlsxwriter
import BCS_connector
from datetime import datetime
import matplotlib.pyplot as plt
import mailer_1
import os



def read_data(id):

    
    qt_df = BCS_connector.pre_reader_orders(id, BCS_connector.open_orders)

    qt_df["job_name"] = qt_df["job_name"].fillna("Not available (blank)")
    qt_df["order_date"] = pd.to_datetime(qt_df["order_date"], errors="coerce")

    print("the coluimns are as follows --- ", qt_df.columns)
    print("Head:  ", qt_df.head())

    return qt_df
    

def remove_all_files_in_folder(folder_path):
    # Loop through all items in the given folder
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        # Check if it's a file and remove it
        if os.path.isfile(file_path):
            os.remove(file_path)
            print(f"Removed: {file_path}")



from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl import load_workbook
def pivot_table(rma_df, prefix, folder_path, new_loop):

    new_dir = "All orders"

    # Combine the folder path with the new directory name
    new_directory_path = os.path.join(folder_path, new_dir)
    os.makedirs(new_directory_path, exist_ok=True)

    if new_loop:
        
        remove_all_files_in_folder(new_directory_path)


    output_file_path = f"{prefix} - Open Order Detail Report.xlsx"

    output_file = os.path.join(new_directory_path, output_file_path)
    
    # Writing to Excel with different sheet names
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        rma_df.to_excel(writer, sheet_name='All open orders', index=False)        


        # Create a pivot table from the rma_df
        rma_pivot_df = pd.pivot_table(
            rma_df, 
            values=['Open_Value'], 
            index=['Created_By', 'Customer_Name', 'po_no', 'job_name'],  
            aggfunc={
                'Open_Value': lambda x: round(np.sum(x), 0)  # Sum and then round
            },
            fill_value=0
        )


        # Create a pivot table from the rma_df
        customer_pivot_df = pd.pivot_table(
            rma_df, 
            values=['Open_Value'], 
            index=['Customer_Name', 'Created_By', 'order_no', 'Salesrep', 'po_no', 'job_name', 'order_date'],  
            aggfunc={
                'Open_Value': lambda x: round(np.sum(x), 0)  # Sum for open_line_value
            },
            fill_value=0
        )

        rma_pivot_df['Open_Value'] = rma_pivot_df['Open_Value'].round(0)
        customer_pivot_df['Open_Value'] = customer_pivot_df['Open_Value'].round(0)

        # Write the pivot table to a new sheet
        customer_pivot_df.to_excel(writer, sheet_name='Pivot Summary - customer basis')
        cus_table = customer_pivot_df.to_html(classes='table table-striped', border=1,  index=False)

        # Write the pivot table to a new sheet
        rma_pivot_df.to_excel(writer, sheet_name='Pivot Summary - Order taker')
        rma_table = rma_pivot_df.to_html(classes='table table-striped', border=1,  index=False)

        # Add custom styling (bold column headers and outer borders)
        rma_table = rma_table.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
        rma_table = rma_table.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
        rma_table = rma_table.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')

        # Add custom styling (bold column headers and outer borders)
        cus_table = cus_table.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
        cus_table = cus_table.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
        cus_table = cus_table.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')


    # Load the workbook to apply formatting
    wb = load_workbook(output_file)
    ws = wb['Pivot Summary - Order taker']

    # Create a border style for the header and title
    border = Border(
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Apply border and bold text to the new title row (first row, originally second)
    for cell in ws[1]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)


    # Ensure no borders or bold font for the data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = None  # Remove borders
            cell.font = Font(bold=False)  # Ensure the font is not bold

    amt_column_index = 5

    # Apply number formatting with commas for the 'amt' column
    for row in ws.iter_rows(min_row=2, min_col=amt_column_index, max_col=amt_column_index):
        for cell in row:
            cell.number_format = '#,##0'  # Comma separator for thousands

    # Set column widths based on the content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook with applied formatting
    wb.save(output_file)



    # Load the workbook to apply formatting
    wb = load_workbook(output_file)
    ws = wb['Pivot Summary - customer basis']

    # Create a border style for the header and title
    border = Border(
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Apply border and bold text to the new title row (first row, originally second)
    for cell in ws[1]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)


    # Ensure no borders or bold font for the data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = None  # Remove borders
            cell.font = Font(bold=False)  # Ensure the font is not bold

    amt_column_index = 8 

    # Apply number formatting with commas for the 'amt' column
    for row in ws.iter_rows(min_row=2, min_col=amt_column_index, max_col=amt_column_index):
        for cell in row:
            cell.number_format = '#,##0'  # Comma separator for thousands

    # Set column widths based on the content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook with applied formatting
    wb.save(output_file)
    

    print(f'Pivot table created and saved to {output_file}')

    df = pd.read_excel(output_file, sheet_name="Pivot Summary - Order taker")
    
    ir_table = df.to_html(classes='table table-striped', border=0,  index=False)
    # Add custom styling (bold column headers and outer borders)
    # Add custom styling (bold column headers and outer borders)
    ir_table = ir_table.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
    ir_table = ir_table.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
    ir_table = ir_table.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')


    return rma_df, new_directory_path


def table_creator(df):

    today = pd.to_datetime("today")

    df["days_open"] = (today - df["order_date"]).dt.days
    
    bins = [0, 30, 60, 90, 180, float('inf')]  # Adjusted bins, the last bin is open-ended
    labels = ['0-30', '30-60', '60-90', '90-180', '180+']
    df['day_range'] = pd.cut(df['days_open'], bins=bins, labels=labels, right=True)

    # Create the pivot table
    pivot_df = pd.pivot_table(
            df, 
            values=['Open_Value'], 
            index=['sales_location', 'day_range'],  
            aggfunc={
                'Open_Value': lambda x: round(np.sum(x), 0)  # Sum and then round
            },
            fill_value=0
        )
    
    pivot_df['Open_Value'] = pivot_df['Open_Value'].round(0).astype(int)
    
    # Convert the pivot table to an HTML table with desired styles
    pivot_html = pivot_df.to_html(classes='table table-striped', border=0, index=True)
    
    # Add custom styling (bold column headers and outer borders)
    # Add custom styling (bold column headers and outer borders)
    pivot_html = pivot_html.replace('<table', '<table style="border-collapse: collapse; border: 2px solid black;"')
    pivot_html = pivot_html.replace('<th>', '<th style="font-weight: bold; text-align: center; padding: 8px; border: 1px solid black;">')
    pivot_html = pivot_html.replace('<td>', '<td style="text-align: center; padding: 8px; border: 1px solid black;">')

    
    return pivot_html




def main():

    folder_path = "D:\\Brian's report automation\\Weekly reports"

    main_df = pd.DataFrame()

    ids = {
        166553: "AUS",  # Austin
        173042: "BOS",  # Boston
        176046: "CA",   # California
        175883: "CHAR", # Charlotte
        166557: "DAL",  # Dallas
        166559: "HOU",  # Houston
        175891: "MIL",  # Milwaukee
        10510: "MIN",   # Minnesota
        175888: "NJ",   # New Jersey
        166560: "NOR",  # New Orleans
        175890: "NY",   # New York
        10006: "PHX",   # Phoenix
        166561: "SAT",  # San Antonio
        10008: "SLC",   # Salt Lake City
        10770: "TN"     # Nashville
    }

    newloop = True

    for id, prefix in ids.items():
        tdf = read_data(str(id))
        print("Read the data from database...")
        rma_df, new_dir_path = pivot_table(tdf, prefix, folder_path, newloop)

        main_df = pd.concat([rma_df, main_df], ignore_index=False)

        newloop = False


    tab_df = table_creator(main_df)


    return new_dir_path, tab_df



if __name__ == "__main__":
    main()

"""
"""